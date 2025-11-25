import openpyxl
from tkinter import Tk, filedialog, Text, Button, Label, END
import os

# -------------------------------
# 初始配置
# -------------------------------
save_dir = r"D:\小云管家电费表格"
download_dir = os.path.join(os.path.expanduser("~"), "Downloads")

# -------------------------------
# 修改保存路径
# -------------------------------
def modify_save_dir():
    global save_dir
    new_dir = filedialog.askdirectory(title="请选择保存文件的文件夹", initialdir=save_dir)
    if new_dir:
        save_dir = new_dir
        label_save_dir.config(text=f"当前另存为路径：{save_dir}")

# -------------------------------
# 文件名过滤非法字符
# -------------------------------
def sanitize_filename(name):
    forbidden_chars = r'\/:*?"<>|'
    for ch in forbidden_chars:
        name = name.replace(ch, '_')
    return name

# -------------------------------
# 处理 Excel 文件
# -------------------------------
def process_file():
    global save_dir
    file_path = filedialog.askopenfilename(
        title="请选择要处理的 Excel 文件",
        initialdir=download_dir,
        filetypes=[("Excel 文件", "*.xlsx")]
    )
    if not file_path:
        return

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 清空 AM、AN、AO 三列
    for col_letter in ['AM', 'AN', 'AO']:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).value = None

    # AL、AP 列从第2行开始转换数字
    for col_letter in ['AL', 'AP']:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                try:
                    cell.value = float(cell.value)
                except ValueError:
                    cell.value = 0

    ws['AN1'] = "计算结果"

    # 计算总金额和总度数
    total_amount = sum(ws.cell(row=row, column=openpyxl.utils.column_index_from_string('AP')).value or 0
                       for row in range(2, ws.max_row + 1))
    total_kwh = sum(ws.cell(row=row, column=openpyxl.utils.column_index_from_string('AL')).value or 0
                    for row in range(2, ws.max_row + 1))

    # AN2、AN3 写公式
    ws['AN2'] = "=SUM(AP:AP)"
    ws['AN3'] = "=SUM(AL:AL)"

    # 构建新文件名 B2在前
    b2_value = ws['B2'].value or "空"
    orig_name = os.path.splitext(os.path.basename(file_path))[0]
    new_filename = f"{sanitize_filename(b2_value)}_{sanitize_filename(orig_name)}.xlsx"
    new_file_path = os.path.join(save_dir, new_filename)

    wb.save(new_file_path)

    # 弹窗显示
    win_result = Tk()
    win_result.title("处理完成")
    win_result.geometry("480x220")
    win_result.configure(bg="#f8f8f8")

    text = Text(win_result, wrap='word', font=("Consolas", 11), bg="#ffffff")
    text.insert(END, f"文件已处理完成并保存为:\n{new_file_path}\n\n总金额: {total_amount}\n总度数: {total_kwh}")
    text.pack(expand=True, fill='both', padx=10, pady=10)
    text.config(state='normal')

    Button(win_result, text="关闭", command=win_result.destroy, font=("Arial", 11), bg="#4CAF50", fg="white").pack(pady=8, ipadx=10, ipady=5)
    win_result.mainloop()

# -------------------------------
# 读取表格内容
# -------------------------------
def read_file():
    file_path = filedialog.askopenfilename(
        title="请选择要读取的 Excel 文件",
        initialdir=r"D:\小云管家电费表格",
        filetypes=[("Excel 文件", "*.xlsx")]
    )
    if not file_path:
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    project_name = ws['B2'].value or "空"

    total_amount = sum(
        ws.cell(row=row, column=openpyxl.utils.column_index_from_string('AP')).value or 0
        for row in range(2, ws.max_row + 1)
    )
    total_kwh = sum(
        ws.cell(row=row, column=openpyxl.utils.column_index_from_string('AL')).value or 0
        for row in range(2, ws.max_row + 1)
    )

    win_result = Tk()
    win_result.title("读取结果")
    win_result.geometry("400x180")
    win_result.configure(bg="#f8f8f8")

    text = Text(win_result, wrap='word', font=("Consolas", 11), bg="#ffffff")
    text.insert(END, f"项目名称: {project_name}\n总金额: {total_amount}\n总度数: {total_kwh}")
    text.pack(expand=True, fill='both', padx=10, pady=10)
    text.config(state='normal')

    Button(win_result, text="关闭", command=win_result.destroy, font=("Arial", 11), bg="#FF9800", fg="white").pack(pady=8, ipadx=10, ipady=5)
    win_result.mainloop()

# -------------------------------
# 主窗口
# -------------------------------
root = Tk()
root.title("Excel 自动处理工具")
root.geometry("520x260")
root.configure(bg="#f8f8f8")

label_save_dir = Label(root, text=f"当前另存为路径：{save_dir}", font=("Arial", 11), bg="#f8f8f8")
label_save_dir.pack(pady=15)

btn_modify = Button(root, text="修改另存为路径", command=modify_save_dir, font=("Arial", 11), bg="#2196F3", fg="white")
btn_modify.pack(pady=6, ipadx=10, ipady=5)

btn_process = Button(root, text="开始处理 Excel 文件", command=process_file, font=("Arial", 11), bg="#4CAF50", fg="white")
btn_process.pack(pady=10, ipadx=10, ipady=5)

btn_read = Button(root, text="读取表格", command=read_file, font=("Arial", 11), bg="#FF9800", fg="white")
btn_read.pack(pady=6, ipadx=10, ipady=5)

root.mainloop()
